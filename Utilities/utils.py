import logging
import softest
from openpyxl import Workbook, load_workbook
import csv

class Utils(softest.TestCase):

    def assertlistitemsText(self, list, value):
        for stop in list:
            print("The text is: ", stop.text)
            self.soft_assert(self.assertEqual(), stop.text, value)
            if stop.text == value:
                print("Test Passed")
            else:
                print("Test Failed")
        self.assert_all()

    def custom_logger(logLevel=logging.DEBUG):

        # Create Logger
        logger = logging.getLogger(__name__)
        logger.setLevel(logging.DEBUG)

        # Create the console handler or file handler
        ch = logging.StreamHandler()

        # Create Format for logging
        formatter = logging.Formatter('%(asctime)s %(levelname)s:%(message)s')
        formatter1 = logging.Formatter('%(asctime)s %(message)s')

        # Add the formatter to console or file handler
        ch.setFormatter(formatter)
        fl = logging.FileHandler(".\\Reports\\automationlogfile.log", "a")

        # Add the console handler to the logger
        logger.addHandler(ch)

        # application code or log messages
        logger.debug("Debug Log Statement")
        logger.info("Info Logging File")
        logger.warning("Warning Logging File")
        logger.error("Error Logging File")
        logger.critical("Critical Logging File")

        return logger

    def read_data_from_excel(file_name, sheet):

        data_list = []

        wb = load_workbook(filename=file_name)
        sh = wb[sheet]

        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct+1):
            row = []
            for j in range(1, col_ct+1):
                row.append(sh.cell(row=1, column=j).value)
            data_list.append(row)

        return data_list


    def read_data_from_csv(filename):
        #create empty list
        data_list = []

        #open csv file
        csv_data = open(filename, "r")

        #read csv data
        reader = csv.reader(csv_data)

        #skip header
        next(reader)

        #add csv data to list

        for rows in reader:
            data_list.append(rows)

        return data_list





